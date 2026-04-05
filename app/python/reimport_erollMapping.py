import psycopg2
import sys
import json
from openpyxl import load_workbook
import csv

def validate_headers(headers, required_columns):
    """Validate if all required columns are present"""
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise Exception(f"Missing required columns: {missing}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python upload_mapping_override.py <excel_file> <data_id>")
        sys.exit(1)

    excel_file = sys.argv[1]
    try:
        DATA_ID = int(sys.argv[2])
    except ValueError:
        print("Error: data_id must be an integer")
        sys.exit(1)

    # Database connection
    conn = psycopg2.connect(
        dbname="tbo_data",
        user="postgres",
        password="admin",
        host="localhost",
        port="5432"
    )

    cur = conn.cursor()

    # Define columns (matching your table structure)
    special_fields = [
        "village_id", "gp_ward_id", "block_id", "psb_id",
        "coordinate_id", "kendra_id", "mandal_id",
        "pjila_id", "pincode_id", "postoff_id", "policst_id"
    ]

    mapping_columns = [
        "data_id", "ac_id", "ac_name", "bhag_no", "bhag", "sec_no", "section", "ru",
        "village", "village_id", "gp_ward", "gp_ward_id", "block", "block_id",
        "psb", "psb_id", "coordinate", "coordinate_id", "kendra", "kendra_id",
        "mandal", "mandal_id", "pjila", "pjila_id", "pincode", "pincode_id",
        "postoff", "postoff_id", "policst", "policst_id"
    ]

    try:
        conn.autocommit = False

        # ===================================================
        # 1️⃣ DELETE OLD MAPPING FOR THIS data_id
        # ===================================================
        print(f"Deleting old mapping for data_id {DATA_ID}...")
        cur.execute("""
            DELETE FROM eroll_mapping
            WHERE data_id = %s
        """, (DATA_ID,))
        deleted_count = cur.rowcount
        print(f"*** Deleted {deleted_count} old records")

        # ===================================================
        # 2️⃣ READ EXCEL/CSV FILE
        # ===================================================
        print(f"Reading file: {excel_file}")
        
        # Check file extension
        if excel_file.endswith('.csv'):
            # Handle CSV file
            with open(excel_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            headers = reader.fieldnames
        else:
            # Handle Excel file
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    headers.append(f"column_{len(headers)}")
            
            # Get data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                    rows.append(row_dict)

        print(f"Found {len(rows)} rows to process")

        # Validate required columns
        required_columns = ["ac_id", "bhag_no", "sec_no"]
        missing_cols = [col for col in required_columns if col not in headers]
        if missing_cols:
            raise Exception(f"Missing required columns: {missing_cols}")

        inserted_count = 0
        updated_eroll_count = 0

        # ===================================================
        # 3️⃣ PROCESS EACH ROW
        # ===================================================
        for row_dict in rows:
            # Skip empty rows
            if not row_dict or all(v is None or v == '' for v in row_dict.values()):
                continue

            # ---------------------------
            # INSERT INTO eroll_mapping
            # ---------------------------
            values = []
            for col in mapping_columns:
                if col == "data_id":
                    values.append(DATA_ID)
                else:
                    val = row_dict.get(col)
                    # Convert empty strings to None
                    if val == '':
                        values.append(None)
                    else:
                        values.append(val)

            placeholders = ", ".join(["%s"] * len(mapping_columns))

            cur.execute(f"""
                INSERT INTO eroll_mapping ({", ".join(mapping_columns)})
                VALUES ({placeholders})
            """, values)
            inserted_count += 1

            # ---------------------------
            # SYNC TO eroll_db
            # ---------------------------
            mapping_json = {}
            for field in special_fields:
                if row_dict.get(field):
                    mapping_json[field] = row_dict.get(field)

            update_parts = []
            update_values = []

            # section sync
            if row_dict.get("section") not in (None, ''):
                update_parts.append("section = %s")
                update_values.append(row_dict.get("section"))

            # bhag_no sync
            if row_dict.get("bhag_no") not in (None, ''):
                update_parts.append("bhag_no = %s")
                update_values.append(int(row_dict.get("bhag_no")))

            # sec_no sync
            if row_dict.get("sec_no") not in (None, ''):
                update_parts.append("sec_no = %s")
                update_values.append(int(row_dict.get("sec_no")))

            # mapping JSONB sync
            if mapping_json:
                update_parts.append("mapping = COALESCE(mapping, '{}'::jsonb) || %s::jsonb")
                update_values.append(json.dumps(mapping_json))

            if update_parts:
                update_parts.append("update_by = CURRENT_TIMESTAMP")

                # Add WHERE clause parameters
                update_values.extend([
                    DATA_ID,
                    int(row_dict.get("ac_id", 0)),
                    int(row_dict.get("bhag_no", 0)),
                    int(row_dict.get("sec_no", 0))
                ])

                cur.execute(f"""
                    UPDATE eroll_db
                    SET {", ".join(update_parts)}
                    WHERE data_id = %s
                    AND ac_no = %s
                    AND bhag_no = %s
                    AND sec_no = %s
                """, update_values)
                
                if cur.rowcount > 0:
                    updated_eroll_count += 1

        # ===================================================
        # 4️⃣ COMMIT TRANSACTION
        # ===================================================
        conn.commit()
        print("\n" + "="*50)
        print("MAPPING OVERRIDE COMPLETED SUCCESSFULLY")
        print("="*50)
        print(f"****Statistics:")
        print(f"   - Records inserted in eroll_mapping: {inserted_count}")
        print(f"   - Records updated in eroll_db: {updated_eroll_count}")
        print(f"   - Total rows processed: {len(rows)}")
        print("="*50)

    except Exception as e:
        conn.rollback()
        print("\n ERROR:", str(e))
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    main()